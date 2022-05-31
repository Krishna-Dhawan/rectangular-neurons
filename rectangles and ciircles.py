import openpyxl as opx
import time
import cv2

img = cv2.imread("ML_Data/Untitled.png", 1)
res = (20, 20)

weights = opx.load_workbook("weights.xlsx")
sheet = weights["Sheet1"]

t = time.time()
bias = 0
o = []
for i in range(20):
    for j in range(20):
        B, G, R = img[i, j]
        o.append((G//255) * sheet.cell(row=(i+1), column=(j+1)).value)

print(sum(o))
print("Time taken:", time.time()-t)

isrect = False
if sum(o) >= bias:
    print("rectangle")
    isrect = True
else:
    print("Not a rectangle")

feedback = input("Am i correct? ")
if feedback == "no":
    for i in range(res[0]):
        for j in range(res[1]):
            B, G, R = img[i, j]
            if G//255 == 1 and isrect:
                sheet.cell(row=(i+1), column=(j+1)).value -= 0.05
            elif G//225 == 1 and not isrect:
                sheet.cell(row=(i+1), column=(j+1)).value += 0.05

weights.save("weights.xlsx")
