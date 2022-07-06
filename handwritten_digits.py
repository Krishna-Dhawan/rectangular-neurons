from __future__ import division
import openpyxl as opx
import cv2
from matrix_multiply import MatrixProduct
import time
import os
import json
from funcns import *

# from sympy import *
# x, y, z, t = symbols('x y z t')
# k, m, n = symbols('k m n', integer=True)
# f, g, h = symbols('f g h', cls=Function)

os.chdir("E:/Documents/GitHub/rectangular-neurons")
tme = time.time()
img = cv2.imread("Untitled1.png", 1)
res = (27, 27)

weights = opx.load_workbook("digits.xlsx")
weights1 = opx.load_workbook("digits1.xlsx")
sheet1 = weights1["Sheet1"]

with open("Activations.json") as fil:
    data = json.load(fil)

M1 = []
M2 = []
for i in range(10):
    """
        Matrix M1 is a 10x(27*27) matrix with all weights of each
        neuron stored in each row. M2 is a column matrix (27*27)x1 
        with px. value / 255 is stored in each row. Output is a 
        10x1 matrix (response1) with activation of each output neuron. 
        The matrix is multiplied with another weight matrix to get  
        final output
    """
    # i = digit, j = row, k = column
    sheet = weights["Sheet" + str(i + 1)]
    L1 = []
    for j in range(res[0]):
        for k in range(res[1]):
            L1.append(sheet.cell(row=(j + 1), column=(k + 1)).value)
            R, G, B = img[j, k]
            M2.append([G / 255])
            # the activations are stored in a json file for later calculations
            data["1"][str(res[0] * j + k)] = G / 255
    M1.append(L1)
response1 = MatrixProduct(M1, M2, 10, (res[0] * res[1]))
print("response1:", response1)
for i in range(len(response1)):
    """
        The activations are turned into a number between 0 
        and 1 using the sigmoid function
    """
    response1[i][0] = sigmoid(response1[i][0])
    data["2"][str(i)] = response1[i][0]

w_list = []
for a in range(10):
    """
        sheet1 contains 10 rows representing 10 weights of
        each of 10 neurons in 10 columns. a is the number
        of digits, w_list is a 1x10 matrix containing weights
        and is multiplied with the 10x10 matrix to give a 1x10 
        matrix representing activation of the digits
    """
    l = []
    for j in range(10):
        l.append(sheet1.cell(row=(a + 1), column=(j + 1)).value)
    w_list.append(l)
response = MatrixProduct(w_list, response1, 10, 10)

with open("Activations.json", "r+") as fil:
    json.dump(data, fil)
    fil.truncate()  # this does something

print("response 1", response1)
# print(response)
out = {}
mx = 0
for i in range(len(response)):
    out[i] = response[i][0]
    if out[i] > out[mx]:
        mx = i
print(out)
print(f"I think the digit you drew is {mx}")
print("time:", time.time() - tme)

feedback = input("am i correct? ")
if feedback == 'no' or feedback == "0":
    ans = int(input("Correct digit: "))

    cost = find_cost(out, ans)
    print(cost)

    for i in range(10):
        a = data["2"][str(i)]
        for j in range(10):
            sh = weights1["Sheet1"]
            w = sh.cell(row=(i + 1), column=(j + 1)).value
            """del C/del w(n, i) = 2 * a(n-1, i)
                        * sigmderiv(w(n, i)*a(n-1, i))
                        * (a(n, i)-y(n, i))
            """
            y = 0
            if i == ans:
                y = 1
            dcbdw = 2 * a * sigmderiv(w * a) * (y - sigmoid(out[i]))
            sh.cell(row=(i + 1), column=(j + 1)).value += dcbdw

    for a in range(10):
        sh = weights["Sheet" + str(a + 1)]
        y_ = 0
        if a == ans:
            y_ = 1
        for i in range(res[0]):
            for j in range(res[1]):
                w = sh.cell(row=(i + 1), column=(j + 1)).value
                ac = data["1"][str(27 * i + j)]
                try:
                    y = y_ / w
                except ZeroDivisionError:
                    y = y_ / 0.01
                # print(w * ac)
                dcbdw = 2 * ac * sigmderiv(w * ac) * (y - data["2"][str(a)])
                sh.cell(row=(i + 1), column=(j + 1)).value += dcbdw

weights.save("digits.xlsx")
weights1.save("digits1.xlsx")
